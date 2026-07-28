"""Exportações de relatórios da carteira processual.

Os arquivos são gerados somente em memória e contêm exclusivamente
processos pertencentes ao usuário autenticado.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from textwrap import shorten
from xml.sax.saxutils import escape

from django.db.models import Q, QuerySet
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from gestao.models import Audiencia, Compromisso, ItemGoogle, LancamentoFinanceiro, MovimentacaoProcesso, Prazo, Processo, PublicacaoDJEN, Tarefa, TriagemJuridica


COLUNAS = [
    ('Número CNJ', 'numero'),
    ('Processo', 'titulo'),
    ('Cliente', 'cliente'),
    ('Área', 'area'),
    ('Status', 'status'),
    ('Tribunal', 'tribunal'),
    ('Vara/Órgão', 'vara'),
    ('Comarca', 'comarca'),
    ('Distribuição', 'data_distribuicao'),
    ('Responsável', 'responsavel'),
    ('Atualização DataJud', 'ultima_sincronizacao_datajud'),
]


def filtrar_processos(user, dados: dict) -> QuerySet[Processo]:
    """Aplica filtros sem jamais ampliar o escopo do usuário."""
    processos = Processo.objects.filter(user=user).select_related('cliente', 'responsavel')
    if cliente := dados.get('cliente'):
        processos = processos.filter(cliente=cliente)
    if area := dados.get('area'):
        processos = processos.filter(area=area)
    if status := dados.get('status'):
        processos = processos.filter(status=status)
    if tribunal := dados.get('tribunal'):
        processos = processos.filter(tribunal__icontains=tribunal.strip())
    if data_inicial := dados.get('data_inicial'):
        processos = processos.filter(data_distribuicao__gte=data_inicial)
    if data_final := dados.get('data_final'):
        processos = processos.filter(data_distribuicao__lte=data_final)
    return processos.order_by('cliente__nome', 'numero', 'titulo')


def descricao_filtros(dados: dict) -> str:
    itens = []
    if cliente := dados.get('cliente'):
        itens.append(f'Cliente: {cliente.nome}')
    if area := dados.get('area'):
        itens.append(f'Área: {dict(Processo.AREA_CHOICES).get(area, area)}')
    if status := dados.get('status'):
        itens.append(f'Status: {dict(Processo.STATUS_CHOICES).get(status, status)}')
    if tribunal := dados.get('tribunal'):
        itens.append(f'Tribunal: {tribunal}')
    if data_inicial := dados.get('data_inicial'):
        itens.append(f'Distribuído de: {data_inicial.strftime("%d/%m/%Y")}')
    if data_final := dados.get('data_final'):
        itens.append(f'Distribuído até: {data_final.strftime("%d/%m/%Y")}')
    return ' | '.join(itens) or 'Sem filtros - todos os processos do usuário'


def _formatar_data(valor) -> str:
    if not valor:
        return '-'
    if isinstance(valor, datetime):
        return timezone.localtime(valor).strftime('%d/%m/%Y %H:%M') if timezone.is_aware(valor) else valor.strftime('%d/%m/%Y %H:%M')
    return valor.strftime('%d/%m/%Y')


def _linhas(processos: QuerySet[Processo]):
    areas = dict(Processo.AREA_CHOICES)
    status = dict(Processo.STATUS_CHOICES)
    for processo in processos:
        atualizacao_datajud = processo.ultima_sincronizacao_datajud
        # O Excel não aceita datetimes com fuso. Mantemos o horário local, sem
        # alterar a informação exibida ao usuário.
        if atualizacao_datajud and timezone.is_aware(atualizacao_datajud):
            atualizacao_datajud = timezone.localtime(atualizacao_datajud).replace(tzinfo=None)
        yield [
            processo.numero or '-', processo.titulo, processo.cliente.nome,
            areas.get(processo.area, processo.area), status.get(processo.status, processo.status),
            processo.tribunal or '-', processo.vara or '-', processo.comarca or '-',
            processo.data_distribuicao, processo.responsavel.get_full_name() if processo.responsavel else '-',
            atualizacao_datajud,
        ]


def gerar_excel_processos(processos: QuerySet[Processo], filtros: dict) -> bytes:
    """Retorna um XLSX legível, com cabeçalho, filtros e datas tipadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Processos'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS))
    ws['A1'] = 'JURI-AI - Relatório de processos'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = PatternFill('solid', fgColor='10243E')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUNAS))
    ws['A2'] = f'Filtros: {descricao_filtros(filtros)}'
    ws['A2'].font = Font(italic=True, color='475569')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COLUNAS))
    ws['A3'] = f'Gerado em {timezone.localtime().strftime("%d/%m/%Y %H:%M")} - {processos.count()} processo(s)'
    ws['A3'].font = Font(size=9, color='64748B')

    header_row = 5
    for index, (titulo, _) in enumerate(COLUNAS, start=1):
        celula = ws.cell(header_row, index, titulo)
        celula.font = Font(bold=True, color='FFFFFF')
        celula.fill = PatternFill('solid', fgColor='1E3A5F')
        celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for linha_index, valores in enumerate(_linhas(processos), start=header_row + 1):
        for coluna, valor in enumerate(valores, start=1):
            celula = ws.cell(linha_index, coluna, valor)
            celula.alignment = Alignment(vertical='top', wrap_text=True)
            if coluna in {9, 11} and valor:
                celula.number_format = 'dd/mm/yyyy hh:mm' if coluna == 11 else 'dd/mm/yyyy'

    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(COLUNAS))}{max(header_row, ws.max_row)}'
    ws.freeze_panes = f'A{header_row + 1}'
    larguras = [21, 32, 28, 17, 16, 16, 23, 20, 15, 24, 23]
    for indice, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 30

    arquivo = BytesIO()
    wb.save(arquivo)
    return arquivo.getvalue()


def gerar_pdf_processos(processos: QuerySet[Processo], filtros: dict) -> bytes:
    """Retorna um PDF horizontal paginado, apropriado para consulta e impressão."""
    arquivo = BytesIO()
    documento = SimpleDocTemplate(
        arquivo, pagesize=landscape(A4), leftMargin=0.75 * cm, rightMargin=0.75 * cm,
        topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        title='JURI-AI - Relatório de processos',
    )
    estilos = getSampleStyleSheet()
    titulo = estilos['Title']
    titulo.textColor = colors.HexColor('#10243E')
    normal = estilos['Normal']
    normal.fontSize = 7
    normal.leading = 9
    pequeno = estilos['BodyText']
    pequeno.fontSize = 8
    pequeno.leading = 10

    dados = [[Paragraph('<b>Número</b>', normal), Paragraph('<b>Processo</b>', normal), Paragraph('<b>Cliente</b>', normal), Paragraph('<b>Área</b>', normal), Paragraph('<b>Status</b>', normal), Paragraph('<b>Tribunal</b>', normal), Paragraph('<b>Distribuição</b>', normal)]]
    for valores in _linhas(processos):
        dados.append([
            Paragraph(shorten(str(valores[0]), width=26, placeholder='...'), normal),
            Paragraph(shorten(str(valores[1]), width=52, placeholder='...'), normal),
            Paragraph(shorten(str(valores[2]), width=35, placeholder='...'), normal),
            Paragraph(str(valores[3]), normal), Paragraph(str(valores[4]), normal),
            Paragraph(shorten(str(valores[5]), width=22, placeholder='...'), normal),
            Paragraph(_formatar_data(valores[8]), normal),
        ])
    if len(dados) == 1:
        dados.append([Paragraph('Nenhum processo encontrado para os filtros selecionados.', pequeno)] + [''] * 6)

    tabela = LongTable(dados, repeatRows=1, colWidths=[3.0 * cm, 6.1 * cm, 4.2 * cm, 2.3 * cm, 2.9 * cm, 2.6 * cm, 2.7 * cm])
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos = [
        Paragraph('JURI-AI - Relatório de processos', titulo),
        Spacer(1, 0.15 * cm),
        Paragraph(f'Filtros: {descricao_filtros(filtros)}', pequeno),
        Paragraph(f'Gerado em {timezone.localtime().strftime("%d/%m/%Y %H:%M")} - {processos.count()} processo(s)', pequeno),
        Spacer(1, 0.35 * cm), tabela,
    ]
    documento.build(elementos)
    return arquivo.getvalue()


def _filtrar_periodo(queryset, campo: str, dados: dict):
    campo_data = f'{campo}__date' if campo.endswith('_hora') or campo in {'criado_em', 'ocorrido_em'} else campo
    if inicio := dados.get('data_inicial'):
        queryset = queryset.filter(**{f'{campo_data}__gte': inicio})
    if final := dados.get('data_final'):
        queryset = queryset.filter(**{f'{campo_data}__lte': final})
    return queryset


def _descricao_filtros_genericos(dados: dict) -> str:
    partes = []
    for campo, rotulo in [('cliente', 'Cliente'), ('prioridade', 'Prioridade'), ('status', 'Status'), ('tipo', 'Tipo'), ('categoria', 'Categoria'), ('fonte', 'Fonte')]:
        valor = dados.get(campo)
        if valor:
            partes.append(f'{rotulo}: {getattr(valor, "nome", valor)}')
    if dados.get('data_inicial'):
        partes.append(f'De: {dados["data_inicial"].strftime("%d/%m/%Y")}')
    if dados.get('data_final'):
        partes.append(f'Até: {dados["data_final"].strftime("%d/%m/%Y")}')
    if dados.get('data'):
        partes.append(f'Data: {dados["data"].strftime("%d/%m/%Y")}')
    return ' | '.join(partes) or 'Sem filtros adicionais'


def gerar_excel_tabela(titulo: str, filtros: str, cabecalhos: list[str], linhas: list[list]) -> bytes:
    """Gera uma planilha padronizada para os relatórios auxiliares."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    total_colunas = len(cabecalhos)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
    ws['A1'] = f'JURI-AI - {titulo}'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = PatternFill('solid', fgColor='10243E')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_colunas)
    ws['A2'] = f'Filtros: {filtros}'
    ws['A2'].font = Font(italic=True, color='475569')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_colunas)
    ws['A3'] = f'Gerado em {timezone.localtime().strftime("%d/%m/%Y %H:%M")} - {len(linhas)} registro(s)'
    ws['A3'].font = Font(size=9, color='64748B')
    for coluna, cabecalho in enumerate(cabecalhos, start=1):
        celula = ws.cell(5, coluna, cabecalho)
        celula.font = Font(bold=True, color='FFFFFF')
        celula.fill = PatternFill('solid', fgColor='1E3A5F')
        celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for linha_numero, valores in enumerate(linhas, start=6):
        for coluna, valor in enumerate(valores, start=1):
            if isinstance(valor, datetime) and timezone.is_aware(valor):
                valor = timezone.localtime(valor).replace(tzinfo=None)
            celula = ws.cell(linha_numero, coluna, valor)
            celula.alignment = Alignment(vertical='top', wrap_text=True)
            if isinstance(valor, datetime):
                celula.number_format = 'dd/mm/yyyy hh:mm'
            elif hasattr(valor, 'strftime') and not isinstance(valor, str):
                celula.number_format = 'dd/mm/yyyy'
    ws.auto_filter.ref = f'A5:{get_column_letter(total_colunas)}{max(5, ws.max_row)}'
    ws.freeze_panes = 'A6'
    for coluna in range(1, total_colunas + 1):
        ws.column_dimensions[get_column_letter(coluna)].width = min(36, max(14, 120 // total_colunas))
    ws.row_dimensions[1].height = 25
    arquivo = BytesIO()
    wb.save(arquivo)
    return arquivo.getvalue()


def gerar_pdf_tabela(titulo: str, filtros: str, cabecalhos: list[str], linhas: list[list]) -> bytes:
    """Gera PDF horizontal paginado com tabela adaptada ao número de colunas."""
    arquivo = BytesIO()
    documento = SimpleDocTemplate(arquivo, pagesize=landscape(A4), leftMargin=0.75 * cm, rightMargin=0.75 * cm, topMargin=0.8 * cm, bottomMargin=0.8 * cm, title=f'JURI-AI - {titulo}')
    estilos = getSampleStyleSheet()
    titulo_estilo = estilos['Title']
    titulo_estilo.textColor = colors.HexColor('#10243E')
    normal = estilos['Normal']
    normal.fontSize, normal.leading = 7, 9
    pequeno = estilos['BodyText']
    pequeno.fontSize, pequeno.leading = 8, 10
    dados_pdf = [[Paragraph(f'<b>{escape(str(item))}</b>', normal) for item in cabecalhos]]
    limite = max(18, 135 // max(1, len(cabecalhos)))
    for linha in linhas:
        dados_pdf.append([Paragraph(escape(shorten(_formatar_data(valor) if isinstance(valor, datetime) or hasattr(valor, 'strftime') else str(valor or '-'), width=limite, placeholder='...')), normal) for valor in linha])
    if len(dados_pdf) == 1:
        dados_pdf.append([Paragraph('Nenhum registro encontrado para os filtros selecionados.', pequeno)] + [''] * (len(cabecalhos) - 1))
    largura_total = 27.6 * cm
    tabela = LongTable(dados_pdf, repeatRows=1, colWidths=[largura_total / len(cabecalhos)] * len(cabecalhos))
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')), ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 3), ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    documento.build([Paragraph(f'JURI-AI - {titulo}', titulo_estilo), Spacer(1, 0.15 * cm), Paragraph(f'Filtros: {escape(filtros)}', pequeno), Paragraph(f'Gerado em {timezone.localtime().strftime("%d/%m/%Y %H:%M")} - {len(linhas)} registro(s)', pequeno), Spacer(1, 0.35 * cm), tabela])
    return arquivo.getvalue()


def dados_relatorio(user, tipo: str, filtros: dict):
    """Retorna título, filtros, colunas e linhas sem incluir dados de outros usuários."""
    cliente = filtros.get('cliente')
    descricao = _descricao_filtros_genericos(filtros)
    if tipo == 'prazos':
        qs = _filtrar_periodo(Prazo.objects.filter(user=user).select_related('processo__cliente', 'responsavel'), 'data_fatal', filtros)
        if cliente: qs = qs.filter(Q(processo__cliente=cliente) | Q(processo__isnull=True, responsavel=user))
        if filtros.get('prioridade'): qs = qs.filter(prioridade=filtros['prioridade'])
        if filtros.get('status'): qs = qs.filter(status=filtros['status'])
        return 'Prazos', descricao, ['Prazo', 'Processo', 'Cliente', 'Termo inicial', 'Data fatal', 'Prioridade', 'Status', 'Conferido'], [[p.titulo, p.processo.titulo if p.processo else '-', p.processo.cliente.nome if p.processo else '-', p.termo_inicial, p.data_fatal, p.get_prioridade_display(), p.get_status_display(), 'Sim' if p.confirmado else 'Não'] for p in qs.order_by('data_fatal')]
    if tipo == 'audiencias':
        qs = _filtrar_periodo(Audiencia.objects.filter(user=user).select_related('processo', 'cliente', 'responsavel'), 'data_hora', filtros)
        if cliente: qs = qs.filter(cliente=cliente)
        if filtros.get('tipo'): qs = qs.filter(tipo=filtros['tipo'])
        if filtros.get('status'): qs = qs.filter(status=filtros['status'])
        return 'Audiências', descricao, ['Data e hora', 'Tipo', 'Status', 'Cliente', 'Processo', 'Local', 'Responsável'], [[a.data_hora, a.get_tipo_display(), a.get_status_display(), a.cliente.nome if a.cliente else '-', a.processo.titulo if a.processo else '-', a.local or '-', a.responsavel.get_full_name() if a.responsavel else '-'] for a in qs.order_by('data_hora')]
    if tipo == 'movimentacoes':
        qs = _filtrar_periodo(TriagemJuridica.objects.filter(user=user, prioridade__in=['ALTA', 'URGENTE']).select_related('processo__cliente'), 'criado_em', filtros)
        if cliente: qs = qs.filter(processo__cliente=cliente)
        if filtros.get('categoria'): qs = qs.filter(categoria=filtros['categoria'])
        if filtros.get('fonte'): qs = qs.filter(fonte=filtros['fonte'])
        return 'Movimentações relevantes', descricao, ['Data', 'Prioridade', 'Categoria', 'Fonte', 'Processo', 'Evento', 'Resumo', 'Providência sugerida'], [[t.criado_em, t.get_prioridade_display(), t.get_categoria_display(), t.get_fonte_display(), t.processo.numero if t.processo else '-', t.titulo_origem, t.resumo or '-', t.providencia_sugerida or '-'] for t in qs.order_by('-criado_em')]
    if tipo == 'financeiro':
        qs = _filtrar_periodo(LancamentoFinanceiro.objects.filter(user=user).select_related('cliente', 'processo'), 'data_vencimento', filtros)
        if cliente: qs = qs.filter(cliente=cliente)
        if filtros.get('tipo'): qs = qs.filter(tipo=filtros['tipo'])
        if filtros.get('status'): qs = qs.filter(status=filtros['status'])
        return 'Financeiro por cliente', descricao, ['Vencimento', 'Tipo', 'Categoria', 'Descrição', 'Cliente', 'Processo', 'Valor (R$)', 'Status', 'Pagamento'], [[l.data_vencimento, l.get_tipo_display(), l.get_categoria_display(), l.descricao, l.cliente.nome if l.cliente else '-', l.processo.numero if l.processo else '-', l.valor, l.get_status_display(), l.data_pagamento] for l in qs.order_by('data_vencimento')]
    if tipo == 'resumo-diario':
        dia = filtros.get('data') or timezone.localdate()
        linhas = []
        djen = PublicacaoDJEN.objects.filter(user=user, data_disponibilizacao=dia).select_related('processo__cliente')
        triagens = TriagemJuridica.objects.filter(user=user, criado_em__date=dia, prioridade__in=['ALTA', 'URGENTE']).select_related('processo__cliente')
        prazos = Prazo.objects.filter(user=user, status='PENDENTE', data_fatal__gte=dia, data_fatal__lte=dia + timedelta(days=10)).select_related('processo__cliente')
        audiencias = Audiencia.objects.filter(user=user, data_hora__date=dia).select_related('processo', 'cliente')
        gmail = ItemGoogle.objects.filter(user=user, fonte='GMAIL', ocorrido_em__date=dia)
        agenda = ItemGoogle.objects.filter(user=user, fonte='AGENDA', ocorrido_em__date=dia)
        tarefas = Tarefa.objects.filter(user=user, status__in=['AFAZER', 'FAZENDO'], prazo__lte=dia).select_related('cliente', 'processo')
        if cliente:
            djen = djen.filter(processo__cliente=cliente); triagens = triagens.filter(processo__cliente=cliente); prazos = prazos.filter(processo__cliente=cliente); audiencias = audiencias.filter(cliente=cliente); tarefas = tarefas.filter(Q(cliente=cliente) | Q(processo__cliente=cliente))
        linhas += [['DJEN', p.data_disponibilizacao, p.processo.numero if p.processo else p.numero_processo or '-', p.processo.cliente.nome if p.processo else '-', p.tipo_comunicacao or 'Publicação', 'Conferir publicação oficial'] for p in djen]
        linhas += [['Movimentação relevante', t.criado_em, t.processo.numero if t.processo else '-', t.processo.cliente.nome if t.processo else '-', t.resumo or t.titulo_origem, t.providencia_sugerida or 'Conferir'] for t in triagens]
        linhas += [['Prazo próximo', p.data_fatal, p.processo.numero if p.processo else '-', p.processo.cliente.nome if p.processo else '-', p.titulo, 'Conferir prazo'] for p in prazos]
        linhas += [['Audiência', a.data_hora, a.processo.numero if a.processo else '-', a.cliente.nome if a.cliente else '-', a.get_tipo_display(), 'Preparar audiência'] for a in audiencias]
        linhas += [['Gmail', g.ocorrido_em, '-', '-', g.titulo, 'Ler e classificar e-mail'] for g in gmail]
        linhas += [['Agenda', g.ocorrido_em, '-', '-', g.titulo, 'Conferir compromisso'] for g in agenda]
        linhas += [['Tarefa pendente', t.prazo, t.processo.numero if t.processo else '-', t.cliente.nome if t.cliente else '-', t.titulo, 'Executar tarefa'] for t in tarefas]
        linhas.sort(key=lambda linha: str(linha[1] or ''))
        return 'Resumo diário consolidado', f'Data-base: {dia.strftime("%d/%m/%Y")}' + (f' | Cliente: {cliente.nome}' if cliente else ''), ['Origem', 'Data', 'Processo', 'Cliente', 'Assunto', 'Providência'], linhas
    raise ValueError('Tipo de relatório inválido.')
